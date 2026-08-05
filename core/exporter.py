"""
导出模块

支持导出Excel和CSV格式，按发运主体分组
"""

import pandas as pd
from pathlib import Path
from typing import Optional, List
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from core.entity_config import EntityConfigManager


class ShippingExporter:
    """发运单导出器"""

    def __init__(self):
        # 导出路径由调用方（用户选择的保存位置）决定，此处不再于当前工作目录
        # 预建无用的 output/ 目录（构造副作用）
        self.output_dir = Path("output")

    def export_to_excel(self, df: pd.DataFrame, file_path: str,
                        project_name: str = '', project_drawing: str = '',
                        construction_no: str = '', author: str = '',
                        reviewer: str = '', total_weight: float = 0) -> str:
        """
        导出到Excel（按发运主体分组）

        Args:
            df: 发运单数据
            file_path: 文件路径
            project_name: 项目名称
            project_drawing: 项目图号
            construction_no: 施工号
            author: 制作人
            reviewer: 审核人

        Returns:
            导出文件路径
        """
        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)

        # 按发运主体分组排序
        grouped_data = self._prepare_grouped_data(df)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # 写入数据（从第6行开始，第5行是表头，不写入pandas表头）
            grouped_data.to_excel(writer, sheet_name='发运单', index=False, startrow=5, header=False)
            ws = writer.sheets['发运单']

            # ========== 定义样式 ==========
            title_font = Font(size=16, bold=True)
            header_font = Font(bold=True)
            info_font = Font(size=10)
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # ========== 第1行：标题（合并居中，带边框）==========
            ws.merge_cells('A1:J1')
            ws['A1'] = '发运单'
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align
            ws['A1'].border = thin_border
            # 给合并区域的所有单元格添加边框
            for col in range(2, 11):
                ws.cell(row=1, column=col).border = thin_border

            # ========== 第2行：发运说明（合并居中，带边框）==========
            ws.merge_cells('A2:J2')
            ws['A2'] = '发运说明：A:散装发运；B:打捆发运；C:装箱发运；D:特殊发运。'
            ws['A2'].font = info_font
            ws['A2'].alignment = center_align
            ws['A2'].border = thin_border
            for col in range(2, 11):
                ws.cell(row=2, column=col).border = thin_border

            # ========== 第3行：项目信息（合并单元格，平分列宽）==========
            # 项目名称：A3-B3合并
            ws.merge_cells('A3:B3')
            ws['A3'] = '项目名称'
            ws['A3'].font = header_font
            ws['A3'].alignment = left_align
            ws['A3'].border = thin_border
            ws['B3'].border = thin_border

            # 项目名称值：C3
            ws['C3'] = project_name
            ws['C3'].alignment = left_align
            ws['C3'].border = thin_border

            # 项目图号：D3
            ws['D3'] = '项目图号'
            ws['D3'].font = header_font
            ws['D3'].alignment = left_align
            ws['D3'].border = thin_border

            # 项目图号值：E3
            ws['E3'] = project_drawing
            ws['E3'].alignment = left_align
            ws['E3'].border = thin_border

            # 施工号：F3
            ws['F3'] = '施工号'
            ws['F3'].font = header_font
            ws['F3'].alignment = left_align
            ws['F3'].border = thin_border

            # 施工号值：G3
            ws['G3'] = construction_no
            ws['G3'].alignment = left_align
            ws['G3'].border = thin_border

            # H3-J3合并为空白
            ws.merge_cells('H3:J3')
            ws['H3'].border = thin_border
            ws['I3'].border = thin_border
            ws['J3'].border = thin_border

            # ========== 第4行：制作/审核信息（合并单元格）==========
            # 制作：A4-B4合并
            ws.merge_cells('A4:B4')
            ws['A4'] = '制作'
            ws['A4'].font = header_font
            ws['A4'].alignment = left_align
            ws['A4'].border = thin_border
            ws['B4'].border = thin_border

            # 制作人：C4
            ws['C4'] = author
            ws['C4'].alignment = left_align
            ws['C4'].border = thin_border

            # 审核：D4
            ws['D4'] = '审核'
            ws['D4'].font = header_font
            ws['D4'].alignment = left_align
            ws['D4'].border = thin_border

            # 审核人：E4
            ws['E4'] = reviewer
            ws['E4'].alignment = left_align
            ws['E4'].border = thin_border

            # 项目总重：F4-G4合并
            ws.merge_cells('F4:G4')
            ws['F4'] = '项目总重'
            ws['F4'].font = header_font
            ws['F4'].alignment = left_align
            ws['F4'].border = thin_border
            ws['G4'].border = thin_border

            # 总重值：H4（整数，无小数点）
            ws['H4'] = f'{total_weight:,.0f} kg'
            ws['H4'].alignment = left_align
            ws['H4'].border = thin_border

            # I4-J4合并为空白
            ws.merge_cells('I4:J4')
            ws['I4'].border = thin_border
            ws['J4'].border = thin_border

            # ========== 第5行：表头（带边框）==========
            headers = ['序号', '物料号', '图号', '规格', '名称', '数量', '净重', '总重(kg)', '发运类型', '备注']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col)
                cell.value = header
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            # ========== 设置数据区域样式 ==========
            data_rows = grouped_data.shape[0]

            for i in range(data_rows):
                row = 6 + i  # 数据从第6行开始
                for col in range(1, 11):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    # 默认自动换行
                    cell.alignment = Alignment(wrap_text=True, vertical='center')

                    # 数字列右对齐
                    if col in [6, 7, 8]:  # 数量、净重、总重
                        cell.alignment = Alignment(horizontal='right', wrap_text=True, vertical='center')

                    # 检查是否是发运主体行（物料号列显示"发运主体"）
                    物料号_cell = ws.cell(row=row, column=2)
                    if 物料号_cell.value == '发运主体':
                        # 发运主体行加粗
                        for c in range(1, 11):
                            ws.cell(row=row, column=c).font = Font(bold=True)
                        # 合并C-D列（发运主体显示）
                        ws.merge_cells(f'C{row}:D{row}')
                        ws.cell(row=row, column=3).border = thin_border
                        ws.cell(row=row, column=4).border = thin_border
                        # 合并E-J列
                        ws.merge_cells(f'E{row}:J{row}')
                        for col in range(5, 11):
                            ws.cell(row=row, column=col).border = thin_border
                    else:
                        # 数据行：总重列使用公式
                        if col == 8:  # 总重(kg)列
                            cell.value = f'=F{row}*G{row}'
                        # 发运类型列居中
                        if col == 9:
                            cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')

            # ========== 调整列宽 ==========
            column_widths = {
                'A': 8,   # 序号
                'B': 20,  # 物料号
                'C': 15,  # 图号
                'D': 18,  # 规格
                'E': 25,  # 名称
                'F': 10,  # 数量
                'G': 10,  # 净重
                'H': 12,  # 总重
                'I': 10,  # 发运类型
                'J': 15,  # 备注
            }
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

        return str(file_path)

    def _prepare_grouped_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        准备分组数据（按发运主体分组，添加分组标题行）

        Args:
            df: 原始发运单数据

        Returns:
            带分组标题的DataFrame
        """
        if df.empty:
            return df

        # 按发运主体排序
        df_sorted = df.copy()

        # 添加分组排序列：01-97按代码排序在前，98、99、00固定在最后
        def get_group_sort(entity):
            if not entity:
                return (2, 99)
            code = entity.split('-')[0] if '-' in entity else entity
            # 处理 '97.0' 格式
            if code.endswith('.0'):
                try:
                    code = str(int(float(code)))
                except ValueError:
                    pass
            try:
                code_num = int(code)
            except ValueError:
                return (2, 99)
            if code_num == 98:
                return (1, 98)
            elif code_num == 99:
                return (1, 99)
            elif code_num == 0:
                return (1, 100)
            else:
                return (0, code_num)

        # 用实体代码作为辅助排序键，确保同一实体的物料排在一起
        def get_entity_code(entity):
            if not entity:
                return ''
            code = entity.split('-')[0] if '-' in entity else entity
            if code.endswith('.0'):
                try: code = str(int(float(code)))
                except: pass
            return code

        # 组内排序：按发运方式（B→A→C→D），再按备注，最后按物料号
        def get_method_sort(method):
            if not method:
                return 5
            method_map = {'B': 1, 'A': 2, 'C': 3, 'D': 4}
            # 取第一个字母（兼容 "B-打捆" 格式）
            code = method[0] if method else ''
            return method_map.get(code, 5)

        # 有备注排在无备注前面：有备注→0，无备注→1
        def get_remark_sort(remark):
            return 0 if remark and str(remark).strip() else 1

        df_sorted['_sort_key'] = df_sorted['发运主体'].apply(get_group_sort)
        df_sorted['_entity_code'] = df_sorted['发运主体'].apply(get_entity_code)
        df_sorted['_method_sort'] = df_sorted['发运类型'].apply(get_method_sort)
        df_sorted['_remark_sort'] = df_sorted['备注'].apply(get_remark_sort)
        df_sorted = df_sorted.sort_values(['_sort_key', '_entity_code', '_method_sort', '_remark_sort', '备注', '物料号'])

        # 构建带分组标题的结果
        result_rows = []
        current_entity = None
        seq_num = 1  # 序号计数器

        for _, row in df_sorted.iterrows():
            entity = row.get('发运主体', '')

            # 如果发运主体变化，添加分组标题行
            if entity != current_entity:
                current_entity = entity
                # 获取发运主体代码和名称
                entity_code = entity.split('-')[0] if '-' in entity else entity
                # 标准化代码（处理 '97.0' 格式）
                if entity_code.endswith('.0'):
                    try:
                        entity_code = str(int(float(entity_code)))
                    except ValueError:
                        pass
                # 使用动态映射获取显示名称
                entity_display = EntityConfigManager.get_entity_display_name(entity_code)

                # 添加分组标题行（C列显示发运主体，D列开始合并）
                title_row = {
                    '序号': seq_num,
                    '物料号': '发运主体',
                    '图号': entity_display,  # C列显示发运主体
                    '规格': '',  # D列开始合并
                    '名称': '',
                    '数量': '',
                    '净重': '',
                    '总重(kg)': '',
                    '发运类型': '',
                    '备注': ''
                }
                result_rows.append(title_row)
                seq_num += 1

            # 添加数据行
            # 处理发运类型，只取字母部分
            shipping_method = row.get('发运类型', '')
            if shipping_method and '-' in shipping_method:
                shipping_method = shipping_method.split('-')[0]
            elif shipping_method and len(shipping_method) > 1:
                shipping_method = shipping_method[0]

            data_row = {
                '序号': seq_num,
                '物料号': row.get('物料号', ''),
                '图号': row.get('图号', ''),
                '规格': row.get('规格', ''),
                '名称': row.get('名称', ''),
                '数量': row.get('数量', 0),
                '净重': row.get('净重', 0),
                '总重(kg)': '',  # 将在Excel中使用公式计算
                '发运类型': shipping_method,
                '备注': row.get('备注', '')
            }
            result_rows.append(data_row)
            seq_num += 1

        return pd.DataFrame(result_rows)

    def export_with_grouping(self, df: pd.DataFrame, file_path: str) -> str:
        """
        带分组标题的Excel导出

        Args:
            df: 发运单数据
            file_path: 文件路径

        Returns:
            导出文件路径
        """
        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)

        # 创建带分组标题的数据
        output_data = []
        current_entity = None

        for _, row in df.iterrows():
            entity = row.get('发运主体', '')

            # 如果发运主体变化，添加分组标题
            if entity != current_entity:
                current_entity = entity
                output_data.append({
                    '序号': '',
                    '物料号': '',
                    '图号': '',
                    '规格': '',
                    '名称': entity,
                    '数量': '',
                    '净重': '',
                    '总重(kg)': '',
                    '发运类型': '',
                    '发运主体': '',
                    '备注': ''
                })

            # 添加数据行
            output_data.append(row.to_dict())

        output_df = pd.DataFrame(output_data)

        # 导出
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            output_df.to_excel(writer, sheet_name='发运单', index=False)

        return str(file_path)

    def export_to_csv(self, df: pd.DataFrame, file_path: str, encoding: str = 'utf-8-sig') -> str:
        """
        导出到CSV

        Args:
            df: 发运单数据
            file_path: 文件路径
            encoding: 编码

        Returns:
            导出文件路径
        """
        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)

        df.to_csv(file_path, index=False, encoding=encoding)
        return str(file_path)

    def export_to_pdf(self, df: pd.DataFrame, file_path: str) -> str:
        """
        导出到PDF（需要安装reportlab）

        Args:
            df: 发运单数据
            file_path: 文件路径

        Returns:
            导出文件路径
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            file_path = Path(file_path)
            file_path.parent.mkdir(parents=True, exist_ok=True)

            doc = SimpleDocTemplate(str(file_path), pagesize=landscape(A4))
            elements = []

            # 标题
            styles = getSampleStyleSheet()
            title = Paragraph('发运单', styles['Title'])
            elements.append(title)

            # 表格数据
            data = [df.columns.tolist()] + df.values.tolist()

            # 创建表格
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            elements.append(table)
            doc.build(elements)

            return str(file_path)

        except ImportError:
            raise ImportError("请安装reportlab: pip install reportlab")
